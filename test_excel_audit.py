import json
from pathlib import Path

from openpyxl import Workbook, load_workbook

from multi_agent_tcp.excel_audit import (
    finalize_service_call_audit,
    format_local_time,
    list_agent_records,
    parse_time_range,
    prepare_service_call_audit,
    query_excel_history,
    render_user_log,
)


def _write_workbook(path: Path, rows: list[list[object]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    for row in rows:
        sheet.append(row)
    workbook.save(path)
    workbook.close()


def _set_cell(path: Path, cell: str, value: object) -> None:
    workbook = load_workbook(path)
    workbook["Sheet1"][cell] = value
    workbook.save(path)
    workbook.close()


def _delete_row(path: Path, row: int) -> None:
    workbook = load_workbook(path)
    workbook["Sheet1"].delete_rows(row)
    workbook.save(path)
    workbook.close()


def _clear_row(path: Path, row: int) -> None:
    workbook = load_workbook(path)
    sheet = workbook["Sheet1"]
    for cell in sheet[row]:
        cell.value = None
    workbook.save(path)
    workbook.close()


def _append_row(path: Path, row: list[object]) -> None:
    workbook = load_workbook(path)
    workbook["Sheet1"].append(row)
    workbook.save(path)
    workbook.close()


def test_excel_audit_writes_agent_user_and_before_after_records(tmp_path: Path) -> None:
    session_dir = tmp_path / "session"
    workbook = tmp_path / "table.xlsx"
    _write_workbook(workbook, [["before"]])

    prepared = prepare_service_call_audit(
        {"session_key": "main+default", "session_dir": str(session_dir), "run_id": "run-1", "source_node_id": "planner"},
        "xltool",
        "run",
        {
            "command": "set-cell",
            "arguments": {"file": str(workbook), "sheet": "Sheet1", "cell": "A1", "value": "after", "in_place": True},
        },
        now=lambda: 1_000.123,
    )
    assert prepared is not None

    _set_cell(workbook, "A1", "after")
    record = finalize_service_call_audit(prepared, {"ok": True, "data": {"changed": True}})

    agent_files = list((session_dir / "excel_ops" / "agent").glob("*.json"))
    user_files = list((session_dir / "excel_ops" / "user").glob("*.md"))
    backups = list((session_dir / "excel_ops" / "backups").glob("*.xlsx"))
    assert len(agent_files) == 1
    assert len(user_files) == 1
    assert backups == []
    assert agent_files[0].name.startswith("1000123_19700101_")
    assert record is not None
    assert record["xltool"]["backupAvailable"] is False
    assert record["xltool"]["beforeAfterStatus"] == "captured"
    assert record["xltool"]["beforeAfter"] == [
        {
            "workbook": str(workbook),
            "sheet": "Sheet1",
            "cell": "A1",
            "row": 1,
            "column": "A",
            "field": "",
            "oldValue": "before",
            "newValue": "after",
        }
    ]
    user_text = user_files[0].read_text(encoding="utf-8")
    assert "xltool.run" in user_text
    assert "Old/new cells: 1" in user_text
    assert "Legacy backup" not in user_text


def test_excel_audit_time_range_render_and_history_query(tmp_path: Path) -> None:
    session_dir = tmp_path / "session"
    workbook = tmp_path / "table.xlsx"
    _write_workbook(workbook, [["before"]])

    prepared = prepare_service_call_audit(
        {"session_key": "main+default", "session_dir": str(session_dir)},
        "xltool",
        "set_cell",
        {"file": str(workbook), "sheet": "Sheet1", "cell": "A1", "value": "after", "in_place": True},
        now=lambda: 1_800_000_000.0,
    )
    assert prepared is not None
    _set_cell(workbook, "A1", "after")
    finalize_service_call_audit(prepared, {"ok": True, "data": {"changed": True}})

    start_ms, end_ms = parse_time_range("2020 1 1 0 0 0 0-2030 1 1 0 0 0 0")
    records = list_agent_records(session_dir, start_ms, end_ms)
    assert len(records) == 1
    assert "set-cell" in render_user_log(session_dir, start_ms, end_ms)

    result = query_excel_history(
        session_dir,
        start_time=format_local_time(start_ms),
        end_time=format_local_time(end_ms),
    )
    assert result["ok"] is True
    assert result["count"] == 1
    assert result["records"][0]["command"] == "set-cell"
    assert result["records"][0]["hasBackup"] is False
    assert result["records"][0]["beforeAfterStatus"] == "captured"
    assert result["records"][0]["beforeAfter"][0]["oldValue"] == "before"
    assert result["records"][0]["beforeAfter"][0]["newValue"] == "after"
    assert result["records"][0]["workbook"] == str(workbook)


def test_excel_audit_history_query_filters_latest_workbook_field_and_category(tmp_path: Path) -> None:
    session_dir = tmp_path / "session"
    workbook_a = tmp_path / "skill-table.xlsx"
    workbook_b = tmp_path / "icon-table.xlsx"
    _write_workbook(workbook_a, [["id", "skill_desc"], [1, "old-a"]])
    _write_workbook(workbook_b, [["old-b"]])

    first = prepare_service_call_audit(
        {"session_key": "main+default", "session_dir": str(session_dir), "source_node_id": "planner"},
        "xltool",
        "set_row",
        {"file": str(workbook_a), "sheet": "Sheet1", "row": 2, "header_row": 1, "values": {"skill_desc": "first"}, "in_place": True},
        now=lambda: 1_800_000_000.0,
    )
    assert first is not None
    _set_cell(workbook_a, "B2", "first")
    finalize_service_call_audit(first, {"ok": True, "data": {"changed": True}})

    second = prepare_service_call_audit(
        {"session_key": "main+default", "session_dir": str(session_dir)},
        "xltool",
        "set_cell",
        {"file": str(workbook_b), "sheet": "Sheet1", "cell": "A1", "value": "second", "in_place": True},
        now=lambda: 1_800_000_100.0,
    )
    assert second is not None
    _set_cell(workbook_b, "A1", "second")
    finalize_service_call_audit(second, {"ok": True, "data": {"changed": True}})

    table_queue = prepare_service_call_audit(
        {"session_key": "main+default", "session_dir": str(session_dir), "script_node_id": "table_queue_service"},
        "table_queue",
        "occupy",
        {"tableNames": ["queue-table.xlsx"]},
        now=lambda: 1_800_000_050.0,
    )
    assert table_queue is not None
    finalize_service_call_audit(table_queue, {"ok": True, "result": {"status": "occupied"}})

    latest = query_excel_history(session_dir, limit=10)
    assert latest["count"] == 2
    assert [item["workbook"] for item in latest["records"]] == [str(workbook_b), str(workbook_a)]

    by_workbook = query_excel_history(session_dir, workbook="skill-table", limit=10)
    assert by_workbook["count"] == 1
    assert by_workbook["records"][0]["workbook"] == str(workbook_a)

    by_field = query_excel_history(session_dir, field="skill_desc", limit=10)
    assert by_field["count"] == 1
    assert by_field["records"][0]["methodName"] == "set_row"
    assert by_field["records"][0]["beforeAfter"][0]["oldValue"] == "old-a"
    assert by_field["records"][0]["beforeAfter"][0]["newValue"] == "first"

    all_records = query_excel_history(session_dir, category="all", limit=10)
    assert all_records["count"] == 3

    queue_records = query_excel_history(session_dir, category="table_queue", workbook="queue-table", limit=10)
    assert queue_records["count"] == 1
    assert queue_records["records"][0]["serviceName"] == "table_queue"


def test_excel_audit_set_cells_records_old_new_by_field(tmp_path: Path) -> None:
    session_dir = tmp_path / "session"
    workbook = tmp_path / "table.xlsx"
    _write_workbook(workbook, [["id", "is_download"], [19507, None]])

    prepared = prepare_service_call_audit(
        {"session_key": "main+default", "session_dir": str(session_dir)},
        "xltool",
        "set_cells",
        {"file": str(workbook), "sheet": "Sheet1", "row": 2, "header_row": 1, "values": {"is_download": 1}, "in_place": True},
        now=lambda: 1_800_000_200.0,
    )
    assert prepared is not None
    _set_cell(workbook, "B2", 1)
    finalize_service_call_audit(prepared, {"ok": True, "data": {"changed": True}})

    result = query_excel_history(session_dir, field="is_download")

    assert result["count"] == 1
    assert result["records"][0]["beforeAfterStatus"] == "captured"
    assert result["records"][0]["beforeAfter"] == [
        {
            "workbook": str(workbook),
            "sheet": "Sheet1",
            "cell": "B2",
            "row": 2,
            "column": "B",
            "field": "is_download",
            "oldValue": "",
            "newValue": 1,
        }
    ]


def test_excel_audit_clear_cell_records_old_new(tmp_path: Path) -> None:
    session_dir = tmp_path / "session"
    workbook = tmp_path / "table.xlsx"
    _write_workbook(workbook, [["id", "name"], [1001, "alpha"]])

    prepared = prepare_service_call_audit(
        {"session_key": "main+default", "session_dir": str(session_dir)},
        "xltool",
        "clear_cell",
        {"file": str(workbook), "sheet": "Sheet1", "cell": "B2", "in_place": True},
        now=lambda: 1_800_000_300.0,
    )
    assert prepared is not None
    _set_cell(workbook, "B2", None)
    finalize_service_call_audit(prepared, {"ok": True, "data": {"changed": True}})

    result = query_excel_history(session_dir)

    assert result["records"][0]["command"] == "clear-cell"
    assert result["records"][0]["beforeAfter"] == [
        {
            "workbook": str(workbook),
            "sheet": "Sheet1",
            "cell": "B2",
            "row": 2,
            "column": "B",
            "field": "",
            "oldValue": "alpha",
            "newValue": "",
        }
    ]


def test_excel_audit_remove_row_records_deleted_row_snapshot(tmp_path: Path) -> None:
    session_dir = tmp_path / "session"
    workbook = tmp_path / "table.xlsx"
    _write_workbook(workbook, [["id", "name", "flag"], [1001, "delete-me", 1], [1002, "keep", 0]])

    prepared = prepare_service_call_audit(
        {"session_key": "main+default", "session_dir": str(session_dir)},
        "xltool",
        "remove_row",
        {"file": str(workbook), "sheet": "Sheet1", "row": 2, "header_row": 1, "in_place": True},
        now=lambda: 1_800_000_400.0,
    )
    assert prepared is not None
    _delete_row(workbook, 2)
    record = finalize_service_call_audit(prepared, {"ok": True, "data": {"changed": True}})

    assert record is not None
    assert record["xltool"]["rowSnapshotStatus"] == "captured_before"
    assert record["xltool"]["beforeRows"][0]["role"] == "deleted"
    assert record["xltool"]["beforeRows"][0]["values"]["id"] == "1001"
    assert record["xltool"]["beforeRows"][0]["values"]["name"] == "delete-me"
    assert record["xltool"]["afterRows"] == []

    result = query_excel_history(session_dir, field="name")
    assert result["count"] == 1
    assert result["records"][0]["beforeRows"][0]["values"]["name"] == "delete-me"


def test_excel_audit_clear_row_records_before_and_after_snapshots(tmp_path: Path) -> None:
    session_dir = tmp_path / "session"
    workbook = tmp_path / "table.xlsx"
    _write_workbook(workbook, [["id", "name", "flag"], [1001, "clear-me", 1]])

    prepared = prepare_service_call_audit(
        {"session_key": "main+default", "session_dir": str(session_dir)},
        "xltool",
        "clear_row",
        {"file": str(workbook), "sheet": "Sheet1", "row": 2, "header_row": 1, "in_place": True},
        now=lambda: 1_800_000_500.0,
    )
    assert prepared is not None
    _clear_row(workbook, 2)
    record = finalize_service_call_audit(prepared, {"ok": True, "data": {"changed": True}})

    assert record is not None
    assert record["xltool"]["rowSnapshotStatus"] == "captured"
    assert record["xltool"]["beforeRows"][0]["values"]["name"] == "clear-me"
    assert record["xltool"]["afterRows"][0]["values"]["name"] == ""


def test_excel_audit_append_row_records_inserted_after_snapshot(tmp_path: Path) -> None:
    session_dir = tmp_path / "session"
    workbook = tmp_path / "table.xlsx"
    _write_workbook(workbook, [["id", "name", "flag"], [1001, "old", 1]])

    prepared = prepare_service_call_audit(
        {"session_key": "main+default", "session_dir": str(session_dir)},
        "xltool",
        "append_row",
        {
            "file": str(workbook),
            "sheet": "Sheet1",
            "header_row": 1,
            "values": {"id": 1002, "name": "inserted", "flag": 0},
            "in_place": True,
        },
        now=lambda: 1_800_000_600.0,
    )
    assert prepared is not None
    _append_row(workbook, [1002, "inserted", 0])
    record = finalize_service_call_audit(prepared, {"ok": True, "data": {"changed": True}})

    assert record is not None
    assert record["xltool"]["rowSnapshotStatus"] == "captured_after"
    assert record["xltool"]["afterRows"][0]["role"] == "inserted"
    assert record["xltool"]["afterRows"][0]["row"] == 3
    assert record["xltool"]["afterRows"][0]["values"]["name"] == "inserted"


def test_excel_history_query_limits_returned_row_snapshots_without_trimming_disk_record(tmp_path: Path) -> None:
    session_dir = tmp_path / "session"
    agent_dir = session_dir / "excel_ops" / "agent"
    agent_dir.mkdir(parents=True)
    rows = [
        {"workbook": "table.xlsx", "sheet": "Sheet1", "row": index, "role": "deleted", "values": {"id": index}}
        for index in range(1, 26)
    ]
    record = {
        "schemaVersion": 1,
        "opId": "many-rows",
        "timestampMs": 1_800_000_700_000,
        "time": "2027-01-15 08:00:00.000",
        "category": "xltool",
        "serviceName": "xltool",
        "methodName": "remove_row",
        "status": "succeeded",
        "arguments": {"file": "table.xlsx", "sheet": "Sheet1", "row": 1},
        "result": {"ok": True, "data": {"changed": True}},
        "xltool": {
            "command": "remove-row",
            "workbook": "table.xlsx",
            "beforeRows": rows,
            "afterRows": [],
            "rowSnapshotStatus": "captured_before",
        },
    }
    record_path = agent_dir / "1800000700000_20270115_080000_000_many-rows.json"
    record_path.write_text(json.dumps(record, ensure_ascii=False), encoding="utf-8")

    result = query_excel_history(session_dir)
    persisted = json.loads(record_path.read_text(encoding="utf-8"))

    assert len(persisted["xltool"]["beforeRows"]) == 25
    assert result["records"][0]["beforeRowsTotal"] == 25
    assert result["records"][0]["beforeRowsReturned"] == 20
    assert result["records"][0]["rowSnapshotTruncated"] is True
    assert "Narrow the workbook" in result["records"][0]["rowSnapshotMessage"]


def test_excel_history_query_maps_legacy_backup_path_to_has_backup(tmp_path: Path) -> None:
    session_dir = tmp_path / "session"
    backup = session_dir / "excel_ops" / "backups" / "legacy.xlsx"
    backup.parent.mkdir(parents=True)
    backup.write_text("old", encoding="utf-8")
    agent_dir = session_dir / "excel_ops" / "agent"
    agent_dir.mkdir(parents=True)
    record = {
        "schemaVersion": 1,
        "opId": "legacy",
        "timestampMs": 1_800_000_000_000,
        "time": "2027-01-15 08:00:00.000",
        "category": "xltool",
        "serviceName": "xltool",
        "methodName": "set_cell",
        "status": "succeeded",
        "arguments": {"file": "legacy.xlsx", "cell": "A1", "value": "new"},
        "result": {"ok": True, "data": {"changed": True}},
        "xltool": {
            "command": "set-cell",
            "workbook": "legacy.xlsx",
            "backupPath": str(backup),
            "targetExisted": True,
        },
    }
    (agent_dir / "1800000000000_20270115_080000_000_legacy.json").write_text(
        json.dumps(record, ensure_ascii=False),
        encoding="utf-8",
    )

    result = query_excel_history(session_dir)

    assert result["count"] == 1
    assert result["records"][0]["hasBackup"] is True
    assert result["records"][0]["backupPath"] == str(backup)
